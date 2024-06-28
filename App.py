import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

def find_matches_and_mark(file):
    # Cargar el libro de trabajo
    wb = load_workbook(file)
    ws_hoja1 = wb['Hoja1']
    ws_hoja2 = wb['Hoja2']

    # Convertir las hojas en dataframes
    df_hoja1 = pd.DataFrame(ws_hoja1.values)
    df_hoja2 = pd.DataFrame(ws_hoja2.values)

    # Asumir que las primeras filas son encabezados
    df_hoja1.columns = df_hoja1.iloc[0]
    df_hoja2.columns = df_hoja2.iloc[0]
    df_hoja1 = df_hoja1[1:]
    df_hoja2 = df_hoja2[1:]

    # Convertir las fechas a un formato comparable
    df_hoja1['FechaConcertacion'] = pd.to_datetime(df_hoja1['FechaConcertacion'], errors='coerce')
    df_hoja2['FechaConcertacion'] = pd.to_datetime(df_hoja2['FechaConcertacion'], errors='coerce')

    # Encontrar coincidencias
    coincidencias = pd.merge(df_hoja1, df_hoja2, on=['Comitente', 'FechaConcertacion'])

    # Resaltar las filas coincidentes en Hoja2
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for index, row in coincidencias.iterrows():
        for i in range(2, ws_hoja2.max_row + 1):  # Asumiendo que hay una fila de encabezado
            if (ws_hoja2.cell(row=i, column=1).value == row['Comitente'] and 
                ws_hoja2.cell(row=i, column=2).value == row['FechaConcertacion']):
                for j in range(1, ws_hoja2.max_column + 1):
                    ws_hoja2.cell(row=i, column=j).fill = fill

    # Guardar el libro de trabajo con las coincidencias resaltadas en un BytesIO
    output = BytesIO()
    wb.save(output)
    return output

st.title("Coincidencias en Hojas de Excel")

# Subir el archivo de Excel
uploaded_file = st.file_uploader("Carga tu archivo de Excel", type=['xlsx'])

if uploaded_file is not None:
    # Procesar el archivo y encontrar coincidencias
    output = find_matches_and_mark(uploaded_file)

    # Descargar el archivo procesado
    st.download_button(
        label="Descargar archivo con coincidencias",
        data=output.getvalue(),
        file_name="resultado_marcado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
